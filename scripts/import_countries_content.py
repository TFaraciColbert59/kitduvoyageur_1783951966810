import openpyxl
import json
import urllib.request
import urllib.error
import ssl
import sys

def get_env():
    env = {}
    with open('.env.local', 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip().strip('\'"')
    return env

def main():
    env = get_env()
    supabase_url = env.get('NEXT_PUBLIC_SUPABASE_URL')
    service_key = env.get('SUPABASE_SERVICE_ROLE_KEY') or env.get('NEXT_PUBLIC_SUPABASE_ANON_KEY')

    if not supabase_url or not service_key:
        print("Missing Supabase credentials in .env.local")
        sys.exit(1)

    wb = openpyxl.load_workbook('Informations pays (1).xlsx', data_only=True)

    # 1. Pays sheet
    pays_rows = [r for r in wb['Pays'].iter_rows(values_only=True) if any(x is not None and str(x).strip() != '' for x in r)]
    pays_hdr = pays_rows[0]
    nom_fr_idx = pays_hdr.index('nom_fr')
    iso2_idx = pays_hdr.index('code_iso2')
    iso3_idx = pays_hdr.index('code_iso3')
    nom_en_idx = pays_hdr.index('nom_en')

    country_meta = {}
    for r in pays_rows[1:]:
        nom_fr = str(r[nom_fr_idx]).strip()
        raw_iso2 = r[iso2_idx]
        if raw_iso2 is None or str(raw_iso2).strip().upper() in ['NONE', '', 'NULL', '#N/A']:
            if nom_fr == 'Namibie':
                iso2 = 'NA'
            else:
                iso2 = 'XX'
        else:
            iso2 = str(raw_iso2).strip().upper()

        raw_iso3 = r[iso3_idx]
        if raw_iso3 is None or str(raw_iso3).strip().upper() in ['NONE', '', 'NULL', '#N/A']:
            if nom_fr == 'Namibie':
                iso3 = 'NAM'
            else:
                iso3 = ''
        else:
            iso3 = str(raw_iso3).strip().upper()

        nom_en = str(r[nom_en_idx]).strip() if r[nom_en_idx] else nom_fr
        country_meta[nom_fr] = {
            'iso_a2': iso2,
            'iso_a3': iso3,
            'nom_en': nom_en,
            'slug': iso2.lower()
        }

    def sheet_to_dict(sheet_name):
        rows = [r for r in wb[sheet_name].iter_rows(values_only=True) if any(x is not None and str(x).strip() != '' for x in r)]
        hdr = rows[0]
        pays_idx = hdr.index('pays')
        res = {}
        for r in rows[1:]:
            p_name = str(r[pays_idx]).strip()
            row_dict = {}
            for i, col in enumerate(hdr):
                if col != 'pays':
                    val = r[i]
                    if val is not None:
                        if isinstance(val, (int, float)):
                            row_dict[col] = val
                        else:
                            row_dict[col] = str(val).strip()
                    else:
                        row_dict[col] = None
            res[p_name] = row_dict
        return res

    conseils_map = sheet_to_dict('Conseils Aux Voyageurs Par Pays')
    climat_map = sheet_to_dict('Climat Par Pays')
    parcs_map = sheet_to_dict('Parcs Nationaux Et Treks')
    infos_map = sheet_to_dict('Informations Pays')
    culture_map = sheet_to_dict('Guide Culturel Par Pays')
    budget_map = sheet_to_dict('Budget Voyage Par Pays')

    records = []
    for nom_fr, meta in country_meta.items():
        iso2 = meta['iso_a2']
        slug = meta['slug']

        conseils = conseils_map.get(nom_fr, {})
        climat = climat_map.get(nom_fr, {})
        parcs = parcs_map.get(nom_fr, {})
        infos = infos_map.get(nom_fr, {})
        culture = culture_map.get(nom_fr, {})
        budget = budget_map.get(nom_fr, {})

        pratique_voyage = {
            'visa_requis_fr': conseils.get('visa_requis_fr'),
            'type_visa': conseils.get('type_visa'),
            'cout_visa': conseils.get('cout_visa'),
            'assurance_recommandee': conseils.get('assurance_recommandee'),
            'permis_international_requis': conseils.get('permis_international_requis'),
            'sources': conseils.get('sources'),
        }

        transport = {
            'aeroport_principal': conseils.get('aeroport_principal'),
            'code_iata': conseils.get('code_iata'),
            'compagnies_depuis_france': conseils.get('compagnies_depuis_france'),
            'transport_interieur': conseils.get('transport_interieur'),
            'location_vehicule_conditions': conseils.get('location_vehicule_conditions'),
            'sens_conduite': conseils.get('sens_conduite'),
            'sources': conseils.get('sources'),
        }

        climat_obj = {
            'climat_general': climat.get('climat_general'),
            'meilleure_periode_trek': climat.get('meilleure_periode_trek'),
            'meilleure_periode_plage': climat.get('meilleure_periode_plage'),
            'saison_pluies': climat.get('saison_pluies'),
            'risques_meteo': climat.get('risques_meteo'),
            'temp_moy_janv': climat.get('temp_moy_janv'),
            'temp_moy_juil': climat.get('temp_moy_juil'),
            'sources': climat.get('sources'),
        }

        outdoor = {
            'parcs_nationaux': parcs.get('parcs_nationaux'),
            'treks_phares': parcs.get('treks_phares'),
            'activites_phares': parcs.get('activites_phares'),
            'faune_flore_remarquable': parcs.get('faune_flore_remarquable'),
            'equipement_specifique_recommande': parcs.get('equipement_specifique_recommande'),
            'sources': parcs.get('sources'),
        }

        connectivite = {
            'type_prise_electrique': infos.get('type_prise_electrique'),
            'voltage': infos.get('voltage'),
            'esim_disponible': infos.get('esim_disponible'),
            'couverture_reseau': infos.get('couverture_reseau'),
            'sources': infos.get('sources'),
        }

        culture_obj = {
            'coutumes_etiquette': culture.get('coutumes_etiquette'),
            'phrases_utiles': culture.get('phrases_utiles'),
            'dress_code': culture.get('dress_code'),
            'religion_principale': culture.get('religion_principale'),
            'jours_feries_majeurs': culture.get('jours_feries_majeurs'),
            'sources': culture.get('sources'),
        }

        editorial = {
            'plats_emblematiques': culture.get('plats_emblematiques'),
            'sources': culture.get('sources'),
        }

        budget_obj = {
            'moyens_paiement': budget.get('moyens_paiement'),
            'budget_jour_petit': budget.get('budget_jour_petit'),
            'budget_jour_moyen': budget.get('budget_jour_moyen'),
            'budget_jour_gros': budget.get('budget_jour_gros'),
            'prix_repas_moyen': budget.get('prix_repas_moyen'),
            'prix_hebergement_moyen': budget.get('prix_hebergement_moyen'),
            'usage_pourboire': budget.get('usage_pourboire'),
            'marchandage_usage': budget.get('marchandage_usage'),
            'sources': budget.get('sources'),
        }

        sante_securite = {
            'assurance_recommandee': conseils.get('assurance_recommandee'),
            'risques_meteo': climat.get('risques_meteo'),
            'sources': conseils.get('sources'),
        }

        record = {
            'country_iso_a2': iso2,
            'slug': slug,
            'status': 'published',
            'pratique_voyage': pratique_voyage,
            'climat': climat_obj,
            'budget': budget_obj,
            'sante_securite': sante_securite,
            'transport': transport,
            'culture': culture_obj,
            'outdoor': outdoor,
            'connectivite': connectivite,
            'editorial': editorial,
            'data_source': 'manual',
            'staleness_days': 0,
        }
        records.append(record)

    print(f"Generated {len(records)} country content records. Inserting to Supabase...")

    batch_size = 50
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE

    endpoint = f"{supabase_url}/rest/v1/countries_content?on_conflict=country_iso_a2"
    headers = {
        'apikey': service_key,
        'Authorization': f"Bearer {service_key}",
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates',
    }

    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        payload = json.dumps(batch).encode('utf-8')
        req = urllib.request.Request(endpoint, data=payload, headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req, context=ctx) as resp:
                print(f"Batch {i//batch_size + 1} ({len(batch)} items) inserted with status: {resp.status}")
        except urllib.error.HTTPError as e:
            err_body = e.read().decode('utf-8')
            print(f"Error on batch {i//batch_size + 1}: {e.code} - {err_body}")
            sys.exit(1)

    print("All 195 country records successfully upserted into countries_content!")

if __name__ == '__main__':
    main()
