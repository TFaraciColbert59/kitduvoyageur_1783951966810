#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================
 BOT IMAGES BIGBUY  –  Le Kit du Voyageur (version Selenium)
 Compatible Python 3.14 — aucune compilation requise
=============================================================
"""

import os, re, sys, json, time, hashlib, logging, unicodedata
from pathlib import Path
from datetime import datetime

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException,
    ElementClickInterceptedException, StaleElementReferenceException,
)
from webdriver_manager.chrome import ChromeDriverManager

from config import *

# ══════════════════════════════════════════════════════════════
#  LOGGING
# ══════════════════════════════════════════════════════════════

def setup_logging():
    Path("logs").mkdir(exist_ok=True)
    fmt = "%(asctime)s [%(levelname)s] %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ]
    )

log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════
#  UTILITAIRES
# ══════════════════════════════════════════════════════════════

def slugify(text: str, max_len: int = 60) -> str:
    if not text:
        return "Inconnu"
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text)
    text = re.sub(r'\s+', "_", text.strip())
    text = re.sub(r'_+', "_", text).strip("_")
    return text[:max_len] or "Inconnu"

def upgrade_to_hd(url: str) -> str:
    url = re.sub(r'_\d{2,4}x\d{2,4}\.', '.', url)
    url = re.sub(r'-\d{2,4}x\d{2,4}\.', '.', url)
    url = re.sub(r'/\d{2,4}x\d{2,4}/', '/', url)
    url = re.sub(r'\?.*$', '', url)
    return url

def load_checkpoint() -> dict:
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_checkpoint(data: dict):
    CHECKPOINT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ══════════════════════════════════════════════════════════════
#  LECTURE EXCEL (openpyxl, pas de pandas)
# ══════════════════════════════════════════════════════════════

def load_products():
    log.info(f"📊 Lecture : {EXCEL_INPUT}")
    if not EXCEL_INPUT.exists():
        log.error(f"❌ Fichier introuvable : {EXCEL_INPUT}")
        sys.exit(1)

    wb = load_workbook(EXCEL_INPUT)
    ws = wb.active
    log.info(f"   Feuille : {ws.title}")

    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    log.info(f"   Colonnes trouvées : {list(headers.keys())}")

    required = [COL_PRODUIT, COL_CAT, COL_SOUS_CAT, COL_SKU, COL_URL]
    missing = [c for c in required if c not in headers]
    if missing:
        log.error(f"❌ Colonnes manquantes : {missing}")
        sys.exit(1)

    products = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_name, col_idx in headers.items():
            row_data[col_name] = ws.cell(row=row_idx, column=col_idx).value

        note = str(row_data.get(COL_NOTE, "") or "").strip()
        if URL_OK_MARKER not in note:
            continue

        url = str(row_data.get(COL_URL, "") or "").strip()
        if not url or not url.startswith("http"):
            continue

        products.append(row_data)

    log.info(f"   → {len(products)} produits à traiter")
    return products

# ══════════════════════════════════════════════════════════════
#  NAVIGATEUR SELENIUM
# ══════════════════════════════════════════════════════════════

def create_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument(f"user-agent={USER_AGENT}")
    options.add_argument("--window-size=1366,768")
    options.add_argument("--lang=fr-FR")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(PAGE_TIMEOUT / 1000)
    return driver

# ══════════════════════════════════════════════════════════════
#  SCRAPER BIGBUY
# ══════════════════════════════════════════════════════════════

EXCLUDE_KEYWORDS = [
    "logo", "icon", "sprite", "banner", "payment", "flag", "pixel",
    "tracking", "1x1", "placeholder", "blank", "arrow", "button",
    "star", "rating", "spinner", "loader", "newsletter", "social",
    "facebook", "twitter", "instagram", "/static/", "/wysiwyg/",
    "catalog/category/",
]

def is_product_image(url: str) -> bool:
    u = url.lower()
    if not re.search(r'\.(jpg|jpeg|png|webp)(\?|$)', u):
        return False
    return not any(kw in u for kw in EXCLUDE_KEYWORDS)

def extract_images(driver) -> list:
    images = set()

    for sel in SELECTORS_IMAGES:
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, sel)
            for el in elements:
                for attr in ["src", "data-src", "data-original", "data-lazy-src",
                             "data-zoom-image", "data-full", "data-large"]:
                    try:
                        u = el.get_attribute(attr)
                    except StaleElementReferenceException:
                        continue
                    if u and u.startswith("http") and is_product_image(u):
                        images.add(upgrade_to_hd(u))
                try:
                    srcset = el.get_attribute("srcset") or ""
                except StaleElementReferenceException:
                    srcset = ""
                for part in srcset.split(","):
                    src = part.strip().split(" ")[0] if part.strip() else ""
                    if src.startswith("http") and is_product_image(src):
                        images.add(upgrade_to_hd(src))
        except Exception:
            pass

    # JSON-LD
    try:
        scripts = driver.find_elements(By.CSS_SELECTOR, 'script[type="application/ld+json"]')
        for script in scripts:
            txt = script.get_attribute("innerHTML") or ""
            for url in re.findall(r'https://[^\s"\']+\.(?:jpg|jpeg|png|webp)', txt, re.I):
                if is_product_image(url):
                    images.add(upgrade_to_hd(url))
    except Exception:
        pass

    # JS variables (Fotorama)
    try:
        js_imgs = driver.execute_script("""
            const r = [];
            document.querySelectorAll('[data-fotorama]').forEach(el => {
                try {
                    const d = JSON.parse(el.getAttribute('data-fotorama') || '{}');
                    (d.data || []).forEach(i => { if(i.img) r.push(i.img); if(i.full) r.push(i.full); });
                } catch(e) {}
            });
            return r;
        """)
        for u in (js_imgs or []):
            if isinstance(u, str) and u.startswith("http") and is_product_image(u):
                images.add(upgrade_to_hd(u))
    except Exception:
        pass

    return list(images)

def get_variants(driver) -> list:
    variants = []
    seen_names = set()

    for sel in SELECTORS_VARIANTS:
        try:
            elems = driver.find_elements(By.CSS_SELECTOR, sel)
            if not elems:
                continue
            for el in elems:
                try:
                    name = (
                        el.get_attribute("aria-label")
                        or el.get_attribute("title")
                        or el.get_attribute("data-option-label")
                        or el.get_attribute("data-value")
                        or (el.text or "").strip()
                        or "Coloris"
                    )
                except StaleElementReferenceException:
                    continue
                name = name.strip()
                if name and name not in seen_names:
                    seen_names.add(name)
                    variants.append({"name": name, "sel": sel, "index": len(variants)})
            if variants:
                break
        except Exception:
            pass

    return variants

def accept_cookies(driver):
    for sel in [
        "#onetrust-accept-btn-handler",
        ".onetrust-accept-btn-handler",
        "#cookieConsentAcceptAll",
        "button[class*='accept'][class*='cookie']",
        "[id*='cookie'] button",
    ]:
        try:
            btns = driver.find_elements(By.CSS_SELECTOR, sel)
            for btn in btns:
                if btn.is_displayed():
                    btn.click()
                    time.sleep(0.6)
                    return
        except Exception:
            pass

def scrape_product(driver, url: str) -> dict:
    result = {"variants": {}, "error": None}

    try:
        log.info(f"   🌐 {url}")
        driver.get(url)
        time.sleep(DELAY_LOAD)
        accept_cookies(driver)

        imgs_default = extract_images(driver)
        log.info(f"      {len(imgs_default)} image(s) en état initial")

        variants = get_variants(driver)

        if not variants:
            log.info("      ℹ️  Pas de variante → dossier 'Principal'")
            if imgs_default:
                result["variants"]["Principal"] = imgs_default
            return result

        log.info(f"      🎨 {len(variants)} variante(s) détectée(s)")

        all_variant_imgs = set()
        for v in variants:
            vname = slugify(v["name"])
            try:
                # Re-récupère les éléments à chaque clic (évite les stale refs)
                elems = driver.find_elements(By.CSS_SELECTOR, v["sel"])
                if v["index"] >= len(elems):
                    continue
                el = elems[v["index"]]
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.3)
                try:
                    el.click()
                except ElementClickInterceptedException:
                    driver.execute_script("arguments[0].click();", el)

                time.sleep(DELAY_VARIANT)
                imgs = extract_images(driver)
                if imgs:
                    result["variants"][vname] = imgs
                    all_variant_imgs.update(imgs)
                    log.info(f"      🎨 {vname} → {len(imgs)} image(s)")
                else:
                    log.warning(f"      ⚠️  Variante {vname} : aucune image")
            except Exception as e:
                log.warning(f"      ⚠️  Clic {vname} : {e}")

        orphans = [u for u in imgs_default if u not in all_variant_imgs]
        if orphans:
            result["variants"]["Non_identifie"] = orphans
            log.info(f"      📁 {len(orphans)} image(s) orpheline(s)")

        if not any(result["variants"].values()):
            log.warning("      ⚠️  Fallback <img> brut...")
            fb = driver.execute_script("""
                return Array.from(document.querySelectorAll('img'))
                    .map(i => i.src || i.dataset.src || '')
                    .filter(s => s.startsWith('http') && /\\.(jpg|jpeg|png|webp)/i.test(s));
            """)
            cleaned = [upgrade_to_hd(u) for u in (fb or []) if is_product_image(u)]
            if cleaned:
                result["variants"]["Principal"] = cleaned

    except TimeoutException:
        result["error"] = f"Timeout ({PAGE_TIMEOUT} ms)"
        log.error(f"      ❌ Timeout")
    except Exception as e:
        result["error"] = str(e)
        log.error(f"      ❌ {e}")

    return result

# ══════════════════════════════════════════════════════════════
#  TÉLÉCHARGEMENT
# ══════════════════════════════════════════════════════════════

def guess_ext(content_type: str, url: str) -> str:
    for ct, ext in [("jpeg", ".jpg"), ("png", ".png"), ("webp", ".webp"), ("gif", ".gif")]:
        if ct in content_type.lower():
            return ext
    m = re.search(r'\.(jpg|jpeg|png|webp|gif)', url.lower())
    return "." + (m.group(1) if m else "jpg")

def download_image(url: str, dest: Path, session: requests.Session, hashes: set) -> str:
    try:
        r = session.get(url, timeout=IMG_TIMEOUT, stream=True)
        r.raise_for_status()

        ct = r.headers.get("Content-Type", "")
        if "image" not in ct and "octet" not in ct:
            return "skip"

        data = b"".join(r.iter_content(65536))
        if len(data) < 1_000:
            return "skip"

        fhash = hashlib.sha256(data).hexdigest()
        if fhash in hashes:
            return "skip"

        ext = guess_ext(ct, url)
        final = dest.with_suffix(ext)
        final.parent.mkdir(parents=True, exist_ok=True)

        with open(final, "wb") as f:
            f.write(data)

        hashes.add(fhash)
        return "ok"

    except Exception:
        return "error"

# ══════════════════════════════════════════════════════════════
#  RAPPORT EXCEL
# ══════════════════════════════════════════════════════════════

def build_report(rapport: list, erreurs: list):
    Path("reports").mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Images"

    headers = [
        "#", "Produit", "Marque", "Catégorie", "Sous-catégorie",
        "SKU BigBuy", "Coloris / Variante", "URL BigBuy",
        "✅ Téléchargées", "⏭️ Doublons", "❌ Erreurs", "Statut"
    ]
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", size=11)

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    colors = {"✅ OK": "E2EFDA", "⚠️ Partiel": "FFF2CC", "❌ ERREUR": "FCE4D6"}

    for row in rapport:
        ws.append([
            row.get("num",""), row.get("produit",""), row.get("marque",""),
            row.get("categorie",""), row.get("sous_categorie",""),
            row.get("sku",""), row.get("coloris",""), row.get("url",""),
            row.get("ok",0), row.get("skip",0), row.get("err",0),
            row.get("statut",""),
        ])
        fc = colors.get(row.get("statut",""))
        if fc:
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = PatternFill("solid", fgColor=fc)

    for col, w in zip("ABCDEFGHIJKL", [5,45,20,25,30,15,22,60,12,12,10,12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(RAPPORT_FILE)
    log.info(f"📋 Rapport : {RAPPORT_FILE}")

    if erreurs:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Erreurs"
        ws2.append(["#", "Produit", "URL", "Type", "Détail"])
        for e in erreurs:
            ws2.append([e.get("num",""), e.get("produit",""),
                        e.get("url",""), e.get("type",""), e.get("detail","")])
        wb2.save(ERREURS_FILE)
        log.info(f"⚠️  Erreurs : {ERREURS_FILE}")

# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    setup_logging()
    log.info("=" * 62)
    log.info("  🚀  BOT IMAGES BIGBUY – Le Kit du Voyageur (Selenium)")
    log.info(f"       {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    log.info("=" * 62)

    for d in ["input","output","logs","reports"]:
        Path(d).mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    products   = load_products()
    checkpoint = load_checkpoint()
    log.info(f"💾 Checkpoint : {len(checkpoint)} produit(s) déjà traité(s)")

    rapport, erreurs = [], []
    hashes = set()

    session = requests.Session()
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Referer"   : "https://www.bigbuy.eu/",
        "Accept"    : "image/webp,image/apng,image/*,*/*;q=0.8",
    })

    log.info("🌐 Lancement du navigateur Chrome...")
    driver = create_driver()

    try:
        total = len(products)

        for idx, row in enumerate(products, 1):
            num      = str(row.get(COL_NUM,"?") or "?").strip()
            produit  = str(row.get(COL_PRODUIT,"Produit") or "Produit").strip()
            cat      = str(row.get(COL_CAT,"Divers") or "Divers").strip()
            sous_cat = str(row.get(COL_SOUS_CAT,"General") or "General").strip()
            sku      = str(row.get(COL_SKU,"") or "").strip()
            url      = str(row.get(COL_URL,"") or "").strip()
            marque   = str(row.get(COL_MARQUE,"") or "").strip()

            log.info(f"\n{'─'*62}")
            log.info(f"[{idx}/{total}] {produit}")
            log.info(f"   {cat} › {sous_cat} | SKU: {sku}")

            ck_key = sku if (sku and sku != "Voir catégorie") else url
            if checkpoint.get(ck_key) == "done":
                log.info("   ✅ Déjà traité – ignoré")
                continue

            base_dir = (
                OUTPUT_DIR
                / slugify(cat)
                / slugify(sous_cat)
                / slugify(f"{produit}_{sku}" if sku and sku != "Voir catégorie" else produit)
            )

            sr = scrape_product(driver, url)

            if sr["error"]:
                erreurs.append({"num":num,"produit":produit,"url":url,
                                 "type":"Scraping","detail":sr["error"]})
                rapport.append({"num":num,"produit":produit,"marque":marque,
                                 "categorie":cat,"sous_categorie":sous_cat,
                                 "sku":sku,"coloris":"—","url":url,
                                 "ok":0,"skip":0,"err":0,"statut":"❌ ERREUR"})
                checkpoint[ck_key] = "error"
                save_checkpoint(checkpoint)
                time.sleep(DELAY_PRODUIT)
                continue

            if not sr["variants"]:
                log.warning("   ⚠️  Aucune image récupérée")
                rapport.append({"num":num,"produit":produit,"marque":marque,
                                 "categorie":cat,"sous_categorie":sous_cat,
                                 "sku":sku,"coloris":"—","url":url,
                                 "ok":0,"skip":0,"err":0,"statut":"⚠️ Partiel"})
                checkpoint[ck_key] = "done"
                save_checkpoint(checkpoint)
                time.sleep(DELAY_PRODUIT)
                continue

            for coloris, img_urls in sr["variants"].items():
                col_dir = base_dir / coloris
                col_dir.mkdir(parents=True, exist_ok=True)

                ok = skip = err = 0
                for i, img_url in enumerate(img_urls, 1):
                    dest   = col_dir / f"{str(i).zfill(2)}"
                    status = download_image(img_url, dest, session, hashes)
                    if   status == "ok"  : ok   += 1
                    elif status == "skip": skip += 1
                    else                 : err  += 1

                statut = "✅ OK" if ok > 0 else ("⚠️ Partiel" if skip > 0 else "❌ ERREUR")
                log.info(f"   📁 {coloris}: {ok} OK | {skip} doublons | {err} erreurs → {statut}")

                rapport.append({
                    "num":num,"produit":produit,"marque":marque,
                    "categorie":cat,"sous_categorie":sous_cat,
                    "sku":sku,"coloris":coloris,"url":url,
                    "ok":ok,"skip":skip,"err":err,"statut":statut,
                })

            checkpoint[ck_key] = "done"
            save_checkpoint(checkpoint)
            time.sleep(DELAY_PRODUIT)

    finally:
        driver.quit()

    log.info(f"\n{'='*62}")
    build_report(rapport, erreurs)

    total_imgs = sum(r.get("ok",0) for r in rapport)
    total_err  = sum(1 for r in rapport if r.get("statut")=="❌ ERREUR")

    log.info(f"\n🏁 TERMINÉ")
    log.info(f"   Images téléchargées : {total_imgs}")
    log.info(f"   Lignes en erreur    : {total_err}")
    log.info(f"   Dossier images      : {OUTPUT_DIR.resolve()}")
    log.info(f"   Rapport Excel       : {RAPPORT_FILE.resolve()}")
    log.info("=" * 62)

if __name__ == "__main__":
    main()